"""
Manager views for the custom admin page at /manage/.
Provides:
  - Login with soundweavers / soundweavers5201314
  - Excel download of all instrument data
  - Excel upload to regenerate .md files (packaged as zip)
"""
import io
import os
import re
import zipfile

from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

MARKDOWN_DIR = os.path.join(settings.BASE_DIR, "content", "instruments")
FRONTMATTER_RE = re.compile(r"^---\s*\n(.*?)\n---", re.DOTALL)

# All frontmatter fields we want in the Excel (order matters)
FM_FIELDS = [
    ("title", "樂器名稱（繁體中文）"),
    ("original_name", "原文名稱（英文）"),
    ("category", "分類"),
    ("country", "來源地區"),
    ("era", "年代"),
    ("image", "圖片網址"),
    ("site_url", "網站連結"),
    ("sound_class", "發聲大類"),
    ("range", "音域"),
    ("instrument_key", "調性"),
    ("hs_class", "H-S 分類"),
    ("family", "家族"),
    ("playing_method", "演奏方式"),
    ("body_listening", "身體聆聽"),
    ("soundscape", "聲音景觀"),
    ("region_type", "區域類型"),
    ("youtube_ids", "YouTube ID"),
]

BODY_KEYS = {
    "introduction": "## 介紹",
    "history": "## 歷史背景",
    "timbre": "## 音色描述",
}

SECTION_ORDER = ["introduction", "history", "timbre"]


def _parse_md(filepath):
    """Parse a markdown file and return (frontmatter_dict, sections_dict)."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    fm_data = {}
    sections = {}

    # Parse frontmatter
    fm_match = FRONTMATTER_RE.match(content)
    if fm_match:
        for line in fm_match.group(1).strip().split("\n"):
            if ":" in line:
                key, _, val = line.partition(":")
                fm_data[key.strip()] = val.strip()
        # Remove frontmatter for body parsing
        body = content[fm_match.end():].strip()
    else:
        body = content.strip()

    # Parse body sections
    for section_key, heading in BODY_KEYS.items():
        pattern = re.compile(
            rf"^{re.escape(heading)}\s*\n(.*?)(?=\n## |\Z)",
            re.DOTALL | re.MULTILINE,
        )
        m = pattern.search(body)
        if m:
            sections[section_key] = m.group(1).strip()
        else:
            sections[section_key] = ""

    return fm_data, sections


def _build_md_content(fm_data, sections):
    """Build .md file content from frontmatter dict and sections dict."""
    lines = ["---"]
    for key, val in fm_data.items():
        if val:
            lines.append(f"{key}: {val}")
    lines.append("---")

    for sk in SECTION_ORDER:
        text = sections.get(sk, "")
        if text:
            heading = BODY_KEYS[sk]
            lines.append(f"\n{heading}\n")
            lines.append(text)

    return "\n".join(lines) + "\n"


# ---- Views ----

def manager_login(request):
    if request.method == "POST":
        username = request.POST.get("username", "")
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("manager_dashboard")
        return render(request, "manager/login.html", {"error": "帳號或密碼錯誤"})
    if request.user.is_authenticated:
        return redirect("manager_dashboard")
    return render(request, "manager/login.html")


def manager_logout_view(request):
    logout(request)
    return redirect("manager_login")


@login_required(login_url="manager_login")
def manager_dashboard(request):
    return render(request, "manager/dashboard.html")


@login_required(login_url="manager_login")
def download_excel(request):
    """Generate and return an Excel file of all instrument data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "樂器總資料庫"

    # Style for header rows
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    sub_font = Font(bold=True, color="666666", size=10)
    sub_fill = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
    link_font = Font(color="1D4ED8", underline="single")

    # Build headers: Row 1 = field keys, Row 2 = Chinese explanations
    all_fields = FM_FIELDS + [
        ("introduction", "介紹內容"),
        ("history", "歷史背景"),
        ("timbre", "音色描述"),
    ]
    field_keys = [f[0] for f in all_fields]

    # Row 1: field keys (bold green header)
    for col_idx, (key, zh) in enumerate(all_fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Row 2: Chinese explanation
    for col_idx, (key, zh) in enumerate(all_fields, 1):
        cell = ws.cell(row=2, column=col_idx, value=zh)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center")

    # Freeze panes so header rows stay visible
    ws.freeze_panes = "A3"

    # Process all markdown files
    if os.path.isdir(MARKDOWN_DIR):
        filenames = sorted(f for f in os.listdir(MARKDOWN_DIR) if f.endswith(".md"))
        for row_idx, filename in enumerate(filenames, 3):
            filepath = os.path.join(MARKDOWN_DIR, filename)
            fm_data, sections = _parse_md(filepath)

            for col_idx, key in enumerate(field_keys, 1):
                if key in fm_data:
                    val = fm_data[key]
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if key == "site_url":
                        cell.font = link_font
                        cell.hyperlink = val
                    elif key == "image":
                        cell.font = link_font
                        cell.hyperlink = val
                elif key in sections:
                    ws.cell(row=row_idx, column=col_idx, value=sections[key])

    # Auto-adjust column widths (capped at 50)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="instruments_database.xlsx"'
    return response


@login_required(login_url="manager_login")
def upload_excel(request):
    """Accept uploaded Excel and return a zip of regenerated .md files."""
    if request.method != "POST":
        return redirect("manager_dashboard")

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        return render(
            request,
            "manager/dashboard.html",
            {"error": "請選擇一個 Excel 檔案上傳"},
        )

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        return render(
            request,
            "manager/dashboard.html",
            {"error": f"無法讀取 Excel 檔案：{e}"},
        )

    # Read headers from row 1 (field keys)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append("")

    # Generate .md files
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue  # Skip empty rows

            fm_data = {}
            sections = {}

            for col_idx, val in enumerate(row):
                if col_idx >= len(headers):
                    break
                key = headers[col_idx]
                if not key or not val:
                    continue
                val_str = str(val).strip()
                if not val_str:
                    continue

                if key in BODY_KEYS:
                    sections[key] = val_str
                else:
                    fm_data[key] = val_str

            title = fm_data.get("title", fm_data.get("original_name", "untitled"))
            original_name = fm_data.get("original_name", title)
            # Generate filename from original_name (slugify)
            filename = _slugify(original_name) + ".md"

            content = _build_md_content(fm_data, sections)
            zf.writestr(filename, content)

    zip_buffer.seek(0)

    response = HttpResponse(
        zip_buffer.getvalue(),
        content_type="application/zip",
    )
    response["Content-Disposition"] = 'attachment; filename="instruments_markdown.zip"'
    return response


def _slugify(name):
    """Convert a name to a safe filename slug."""
    name = name.lower().strip()
    # Replace common separators
    name = name.replace(" ", "-").replace("_", "-")
    # Remove non-ascii, non-alphanumeric, non-hyphen chars
    name = re.sub(r"[^a-z0-9-]", "", name)
    # Collapse multiple hyphens
    name = re.sub(r"-+", "-", name)
    # Strip leading/trailing hyphens
    name = name.strip("-")
    return name or "instrument"
